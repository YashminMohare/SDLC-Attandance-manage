from flask import Flask, request, render_template, send_file, jsonify
from PIL import Image
import pytesseract
import openpyxl
import pandas as pd
import io

app = Flask(__name__)
app.secret_key = 'secret_key'  # Secret key for session management, if needed

# Route for the index page
@app.route('/')
def index():
    return render_template('index.html')  # Renders the HTML form for file uploads

# Route to process uploaded files
@app.route('/process', methods=['POST'])
def process_files():
    # Retrieve the uploaded medical certificates and Excel file
    medical_certificates = request.files.getlist('medical_certificate') 
    excel_file = request.files['excel_file']
    
    students_data = []  # List to store extracted student data from certificates
    
    # Process each medical certificate
    for medical_certificate in medical_certificates:
        image = Image.open(medical_certificate)  # Open the image file
        text = pytesseract.image_to_string(image)  # Extract text using OCR (Tesseract)

        student_name = ''
        enrollment_number = ''
        
        # Parse student name and enrollment number from the extracted text
        for line in text.split('\n'):
            if 'Name:' in line:
                student_name = line.split(':')[1].strip()  # Extract name after 'Name:'
            elif 'Enrollment Number:' in line:
                enrollment_number = line.split(':')[1].strip()  # Extract enrollment number after 'Enrollment Number:'

        # Append student data if both name and enrollment number are found
        if student_name and enrollment_number:
            students_data.append({'name': student_name, 'enrollment_number': enrollment_number})

    # Load the uploaded Excel file into an openpyxl workbook
    wb = openpyxl.load_workbook(excel_file)
    sheet = wb.active  # Access the active sheet
    
    # Find the 'Total Attendance' column
    total_attendance_col = None
    for col in range(1, sheet.max_column + 1):
        if sheet.cell(row=1, column=col).value == 'Total Attendance':
            total_attendance_col = col
            break

    # If the 'Total Attendance' column is not found, return an error response
    if total_attendance_col is None:
        return jsonify({'error': 'Total Attendance column not found'}), 400

    # Insert a new column for 'Medical Certificate' status next to 'Total Attendance'
    medical_certificate_col = total_attendance_col + 1
    sheet.insert_cols(medical_certificate_col)
    sheet.cell(row=1, column=medical_certificate_col).value = 'Medical Certificate'  # Header for the new column

    # Loop through each row in the Excel sheet and update the 'Medical Certificate' column
    for row in range(2, sheet.max_row + 1):
        row_student_name = sheet.cell(row=row, column=2).value  # Student name in column 2
        row_enrollment_number = sheet.cell(row=row, column=1).value  # Enrollment number in column 1
        
        matched = False
        # Match the student data with extracted names and enrollment numbers
        for student in students_data:
            if row_student_name == student['name'] or row_enrollment_number == student['enrollment_number']:
                sheet.cell(row=row, column=medical_certificate_col).value = 'Yes'  # Mark 'Yes' for matched students
                matched = True
                break

        if not matched:
            sheet.cell(row=row, column=medical_certificate_col).value = ''  # Leave blank if no match

    # Save the updated workbook into an in-memory output stream
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    # Load the updated workbook as a pandas DataFrame for further processing
    df = pd.read_excel(output)

    # Ensure the necessary columns exist in the Excel file
    if 'Total Attendance' not in df.columns or 'Medical Certificate' not in df.columns:
        return "Required columns are missing in the uploaded file.", 400

    # Add a new 'Updated Attendance' column as a copy of 'Total Attendance'
    df['Updated Attendance'] = df['Total Attendance']  

    # Update attendance for students with medical certificates
    for index, row in df.iterrows():
        if row['Total Attendance'] < 60:  # Check if attendance is below 60%
            if row['Medical Certificate'] == 'Yes':  # Check if medical certificate is present
                df.at[index, 'Updated Attendance'] += 5  # Add 5 points to the attendance

    # Save the final updated DataFrame into a new in-memory output stream
    final_output = io.BytesIO()
    df.to_excel(final_output, index=False)
    final_output.seek(0)

    # Send the updated Excel file back as a downloadable attachment
    return send_file(final_output, as_attachment=True, download_name='updated_attendance.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

if __name__ == '__main__':
    app.run(debug=True)  #run the app in debug mode
